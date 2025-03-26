from openpyxl import load_workbook, Workbook
import pandas as pd
import configparser

# Загружаем конфигурацию из connect_domain.ini
config = configparser.ConfigParser(interpolation=None)
config.read('connect_domain.ini', encoding='utf-8')
MODE = config.get('SETTINGS', 'mode', fallback='new')


# Подключение поиска
from outher.search import user_verification as search_user_verification
from connect.SQLConnect import DatabaseConnector
db = DatabaseConnector() if MODE == "new" else None  # Подключаем БД только если режим "new"

# подключение файла сообщений
from message.message import log

# Подключение Person
from outher.person import Person
from outher.encryption import encrypt_inn

# Подключение BitrixConnect
from connect.bitrixConnect import Bitrix24Connector
bitrix_connector = Bitrix24Connector()
bx24, tokens = bitrix_connector.connect()


# Подключение connect1c
from connect.connect1C import Connector1C
connector_1c = Connector1C()

# Подключение ldapConnect
from connect.ldapConnect import ActiveDirectoryConnector
connector = ActiveDirectoryConnector()
base_dn = connector.getBaseDn()
state = connector.getState()
name_domain = connector.domain_name

# Подключение к SM
from connect.SMConnect import SMConnect

# Выбор версии user_verification
def user_verification(*args):
    """
    Автоматически выбирает нужную версию функции:
    - Если mode = old → Pandas-версия
    - Если mode = new → PostgreSQL-версия
    """
    if MODE == "old":
        df_roles, df_users = args
        return search_user_verification(df_roles, df_users)  # Используем старую версию
    else:
        department, position = args
        return db.user_verification(department=department, position=position)  # Используем новую версию



# Функция записи логина и пароля
def save_login(phone_number, full_name, login):
        bitrix_connector.send_msg_adm(f"{phone_number} {full_name} {login}")
        file_path = 'logins.xlsx'
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Номер телефона', 'ФИО', 'Логин'])
        sheet.append([phone_number, full_name, login])
        workbook.save(file_path)
        workbook.close()

def create_user(file_path):
    global base_dn, state, name_domain, base_dn

    userData = load_workbook(file_path).active

    # поиск по info.xlsx
    if MODE == "old":
        df_users = pd.read_excel(file_path)
        df_roles = pd.read_excel(connector.dbinfo)
        flags = user_verification(df_roles, df_users)  # Используем Pandas
    else:
        department = userData['G2'].value
        position = userData['J2'].value
        flags = user_verification(department, position)  # Используем PostgreSQL

    # Создание объекта сотрудника
    employee = Person(userData['C2'].value, userData['B2'].value, userData["D2"].value)
    phone = userData['K2'].value if not (userData['K2'].value is None) else ''

    # Зашифровка ИНН
    INN = encrypt_inn(userData['A2'].value)

    bitrix_connector.send_msg(str(flags))
    # Cоздание в Active Directory
    ad_success = True
    if flags['AD'] and flags['Normal_account']:
        existence = connector.search_in_ad(INN)

        if existence is None or len(existence) == 0:
            personal = connector.search_by_fullname(employee.full_name)
            simple_email = connector.search_by_mail(employee.create_email(employee.simple_login))
            if personal is None or len(personal) == 0 and simple_email is None or len(simple_email) == 0:
                try:
                    ad_success = connector.create_user(employee.simple_login, employee, userData, INN)
                except Exception as e:
                    bitrix_connector.send_msg_error(
                        f"AD. Создание: Ошибка при создании первичного логина у сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value}.Ошибка {e}")
            else:
                long_email = connector.search_by_mail(employee.create_email(employee.long_login))
                full_email = connector.search_by_mail(employee.create_email(employee.full_login))
                if long_email is None or len(long_email) == 0:
                    try:
                        ad_success = connector.create_user(employee.long_login, employee, userData, INN)
                    except Exception as e:
                        ad_success = False
                        bitrix_connector.send_msg_error(
                            f"AD. Создание: Ошибка при создании вторичного логина у сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value}.Ошибка {e}")
                elif full_email is None or len(full_email) == 0:
                    try:
                        ad_success = connector.create_user(employee.full_login, employee, userData, INN)
                    except Exception as e:
                        ad_success = False
                        bitrix_connector.send_msg_error(
                            f"AD. Создание: Ошибка при создании третичного логина у сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value}.Ошибка {e}")
                else:
                    bitrix_connector.send_msg_error(f"AD.Создание: У сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value}. Поиск по mail выдал, что такой пользователь уже существует в AD")
        else:
            user_dn, attributes = existence[0]
            user_account_control = attributes.get('userAccountControl', [b''])[0]
            uac_value = int(user_account_control.decode())
            is_active = not (uac_value & 0x0002)
            if not is_active:
                try:
                    ad_success = connector.activate_user(user_dn, employee, userData)
                except Exception as e:
                    ad_success = False
                    bitrix_connector.send_msg_error(
                        f"AD. Ошибка при активации учетной записи сотрудника {employee.firstname} {employee.lastname} {employee.surname}: {e}")
            else:
                bitrix_connector.send_msg(
                    f"AD. Создание: У сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value}. Пользователь уже активен в AD {user_dn}.")

    # Создание в Bitrix24
    bx24_success = True
    if flags['AD'] and flags['BX24'] and flags['Normal_account']:
        existence = connector.search_in_ad(INN)
        if not (existence is None) and len(existence) > 0:
            user_dn, attributes = existence[0]
            email = attributes.get('mail', [b''])[0].decode('utf-8')
            user_info = bitrix_connector.search_email(email)

            if user_info == []:
                try:
                    bx24_success = bitrix_connector.create_user(email, employee, userData)
                except Exception as e:
                    bx24_success = False
                    bitrix_connector.send_msg_error(
                        f"BX24. Создание: Ошибка при создании первичного логина у сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value} Ошибка {e}")
            if user_info:
                new_data = {
                    "NAME": userData['C2'].value,
                    "LAST_NAME": userData['B2'].value,
#                    "UF_DEPARTMENT": userData['H2'].value,
                    "UF_DEPARTMENT": "920",
                    "ACTIVE": "Y",
                    "WORK_POSITION": userData['J2'].value
                }
                try:
#                    bx24_success = bitrix_connector.update_user(user_info.get('ID'), new_data, employee, userData)
                    bx24_success, update_status_bx  = bitrix_connector.update_user(user_info, new_data, employee, userData)
                except Exception as e:
                    bx24_success = False
                    bitrix_connector.send_msg_error(
                        f"BX24. Создание: Ошибка при обновлении данных сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value} {user_info.get('ID')} {new_data} Ошибка {e}")
        else:
            bitrix_connector.send_msg_error(
                f"BX24. Создание: У сотрудника {employee.firstname} {employee.lastname} {employee.surname} из отдела {userData['G2'].value} на должность {userData['J2'].value}. Пользователь не найден в AD")

    # Создание в 1С
    c1_success = True
    if flags['ZUP'] or flags['RTL'] or flags['ERP'] and flags['Normal_account']:
        action = "Создание"
        existence = connector.search_in_ad(INN)
        if not (existence is None) and len(existence) > 0:
            user_dn, attributes = existence[0]
            cn = attributes.get('cn', [b''])[0].decode('utf-8')
            login = f"\\BINLTD\{cn}"
            friendly = bitrix_connector.find_jobfriend(userData['J2'].value, userData['H2'].value)
            ZUP_value, RTL_value, ERP_value = (1 if flags['ZUP'] else 0, 1 if flags['RTL'] else 0, 1 if flags['ERP'] else 0)
            url = connector_1c.getUrlCreate()
            data = {
                'full_name': f"{employee.lastname} {employee.firstname} {employee.surname}",
                'password': 'qwerty32',
                'domain': login,
                'ERP': ERP_value,
                'RTL': RTL_value,
                'ZUP': ZUP_value,
                'job_friend': friendly
            }
            c1_success = connector_1c.send_rq(url, data, employee, userData,action)

    # Создание в SM Глобальном
    sm_success = True
    if flags['SM_GEN'] and flags['Normal_account']:
        existence = connector.search_in_ad(INN)
        if existence:
            user_dn, attributes = existence[0]
            login = attributes.get('sAMAccountName', [b''])[0].decode('utf-8')
            sm_login = employee.transform_login(login)
            sm_conn = SMConnect()
            sm_conn.connect_SM()
            test_role_id = sm_conn.getRoleID()
            user_not_exists = sm_conn.user_exists(sm_login) == -1
            if user_not_exists:
                try:
                    sm_success = sm_conn.create_user(sm_login, employee.password, test_role_id)
                    bitrix_connector.send_msg(
                        f"СуперМаг Глобальный. Создание: Сотрудник {employee.firstname} {employee.lastname} {employee.surname} ({sm_login}). Выполнено")
                except Exception as e:
                    sm_success = False
                    bitrix_connector.send_msg_error(
                        f"СуперМаг Глобальный. Создание: Ошибка при создании логина в SM для сотрудника {employee.firstname} {employee.lastname} {employee.surname} ({sm_login}). Ошибка: {e}")
            else:
                bitrix_connector.send_msg(
                    f"СуперМаг Глобальный. Создание: У сотрудника {employee.firstname} {employee.lastname} {employee.surname} логин {sm_login} уже существует.")

    # Получение название локальной базы
    def get_storeId():
        stores_id = userData['N2'].value
        if stores_id is None:
            sm_local_success = False
            return sm_local_success

        if isinstance(stores_id, str):
            stores = [int(num.strip()) for num in stores_id.split(',') if num.strip().isdigit()]
            store_names = []
            for store in stores:
                name_store = sm_conn.get_store(store)
                if name_store:
                    store_names.append(name_store)
            return store_names
        return []

    sm_conn = SMConnect()
    sm_conn.connect_SM()
    store_names = get_storeId()
    sm_local_success = True
    if flags['SM_LOCAL'] and store_names:
        existence = connector.search_in_ad(INN)
        created_databases = []
        user_dn, attributes = existence[0]
        login = attributes.get('sAMAccountName', [b''])[0].decode('utf-8')
        sm_login = employee.transform_login(login)


        for dbname in store_names:
            sm_conn_local = SMConnect()
            sm_conn_local.connect_SM_LOCAL(dbname['dbname'])
            test_role_id = sm_conn_local.getRoleID()

            user_not_exists = sm_conn_local.user_exists(sm_login) == -1
            if user_not_exists:
                try:
                    success = sm_conn_local.create_user_in_local_db(dbname, sm_login, employee.password, test_role_id)
                    sm_local_success = sm_local_success and success
                    if success:
                        created_databases.append(dbname)
                    else:
                        bitrix_connector.send_msg_error(
                            f"СуперМаг Локальный. Ошибка при создании аккаунта для {employee.lastname} {employee.firstname} {employee.surname} в базе {dbname['dbname']}"
                        )
                except Exception as e:
                    sm_local_success = False
                    bitrix_connector.send_msg_error(
                            f"СуперМаг Локальный. Ошибка при создании аккаунта для {employee.lastname} {employee.firstname} {employee.surname} в базе {dbname['dbname']}"
                        )
            else:
                bitrix_connector.send_msg(
                    f"СуперМаг Локальный. Создание: У сотрудника {employee.firstname} {employee.lastname} {employee.surname} в базе {dbname['dbname']} логин {sm_login} уже существует.")
        if created_databases:
            bitrix_connector.send_msg(
                f"СуперМаг Локальный. Создание: Сотруднику {employee.lastname} {employee.firstname} {employee.surname} "
                f"на должность {userData['J2'].value} созданы аккаунты в базах: {', '.join(created_databases)} "
                f"с логином {sm_login}"
            )

    if ad_success and bx24_success and c1_success and sm_success and sm_local_success:
        return True
    else:
        return False

